from openpyxl import *
import statistics
from scipy import stats
import imgkit
import sys
import math
import os

import pprint
pp = pprint.PrettyPrinter(indent=4)


# Build Data Structure
# Parse Wellness_Quotient.xlsx & fill the Data Stucture with Personal Info of Respondent.
# Parse QuizResponses.xlsx & calculate WQ Score & fill remaining calculated-data in the Data Structure.
# Create Output files.

# The Uber Data Structure WQ_Data. Empty Summary is created now.
# Individual Respondent's Data Structure is created later
WQ_Data = {"Summary": {"Overall" : {"N_Respondents": 0,
                                    "N_Males"      : 0,
                                    "N_Females"    : 0,
                                    "N_Cities"     : 0,
                                    "Avg_Form_Time": 0,
                                    "Avg_WQ_Score" : 0,
                                    "Median"       : 0,
                                    "Avg_BMI"      : 0
									},
					   "Cat1_0_18": {"N_Respondents": 0,
                                    "N_Males"      : 0,
                                    "N_Females"    : 0,
                                    "N_Cities"     : 0,
                                    "Avg_Form_Time": 0,
                                    "Avg_WQ_Score" : 0,
                                    "Median"       : 0,
                                    "Avg_BMI"      : 0
									},
					   "Cat2_18_24": {"N_Respondents": 0,
                                    "N_Males"      : 0,
                                    "N_Females"    : 0,
                                    "N_Cities"     : 0,
                                    "Avg_Form_Time": 0,
                                    "Avg_WQ_Score" : 0,
                                    "Median"       : 0,
                                    "Avg_BMI"      : 0
									},
					   "Cat3_25_34": {"N_Respondents": 0,
                                    "N_Males"      : 0,
                                    "N_Females"    : 0,
                                    "N_Cities"     : 0,
                                    "Avg_Form_Time": 0,
                                    "Avg_WQ_Score" : 0,
                                    "Median"       : 0,
                                    "Avg_BMI"      : 0
									},
					   "Cat4_35_44": {"N_Respondents": 0,
                                    "N_Males"      : 0,
                                    "N_Females"    : 0,
                                    "N_Cities"     : 0,
                                    "Avg_Form_Time": 0,
                                    "Avg_WQ_Score" : 0,
                                    "Median"       : 0,
                                    "Avg_BMI"      : 0
									},
					   "Cat5_45_54": {"N_Respondents": 0,
                                    "N_Males"      : 0,
                                    "N_Females"    : 0,
                                    "N_Cities"     : 0,
                                    "Avg_Form_Time": 0,
                                    "Avg_WQ_Score" : 0,
                                    "Median"       : 0,
                                    "Avg_BMI"      : 0
									},
					   "Cat6_55_64": {"N_Respondents": 0,
                                    "N_Males"      : 0,
                                    "N_Females"    : 0,
                                    "N_Cities"     : 0,
                                    "Avg_Form_Time": 0,
                                    "Avg_WQ_Score" : 0,
                                    "Median"       : 0,
                                    "Avg_BMI"      : 0
									},
					   "Cat7_65_120": {"N_Respondents": 0,
                                    "N_Males"      : 0,
                                    "N_Females"    : 0,
                                    "N_Cities"     : 0,
                                    "Avg_Form_Time": 0,
                                    "Avg_WQ_Score" : 0,
                                    "Median"       : 0,
                                    "Avg_BMI"      : 0
									}
						}
			}

# Read individual respondent's response.
WQ_QR = load_workbook('Wellness_Quotient.xlsx')
WQ_QR_sheet = WQ_QR.active

# Defining a global var to use inside varied blocks of scope.
Resp_ID = 0

# Fill the Uber Data Structure WQ_Data with individual Respondent's Data
for WQ_QR_row in WQ_QR_sheet.iter_rows(1, WQ_QR_sheet.max_row,1,159): # Fetching all columns
	for WQ_QR_row_cell in WQ_QR_row:
		# Handle 1st row appropriately
		if WQ_QR_row_cell.row == 1:
			if WQ_QR_row_cell.value != None:
				last_correct_cell_value = WQ_QR_row_cell.value
			else:
				WQ_QR_row_cell.value = last_correct_cell_value
		elif WQ_QR_row_cell.row == 2: # Handle 2nd row appropriately
			if WQ_QR_row_cell.value == None:
				WQ_QR_row_cell.value = WQ_QR_sheet.cell(1, WQ_QR_row_cell.column).value
		else:
			# Handle each column value appropriately
			if WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'Respondent ID':
				Resp_ID = WQ_QR_row_cell.value
				Temp_Dict = {Resp_ID: {"Name"   : "",
                                       "Mobile" : "",
                                       "Email"  : "",
                                       "S_Date" : "",
                                       "E_Date" : "",
                                       "RespTime": "",
                                       "City"   : "",
                                       "AgeCat" : "",
                                       "Gender" : "",
                                       "HeightCm":0,
                                       "WeightKg":0,
                                       "BMI"    : 0,
                                       "Temp_Tot":0,
                                       "Score"  : 0,
                                       "O_Perctl":0,
                                       "AG_Perctl":0,
                                       "AuthCode": ""
									   }
							 }
				# Increase number of overall Respondents in Summary
				WQ_Data["Summary"]["Overall"]["N_Respondents"] += 1
				WQ_Data.update(Temp_Dict)
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'Start Date':
				WQ_Data[Resp_ID]["S_Date"] = WQ_QR_row_cell.value
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'End Date':
				WQ_Data[Resp_ID]["E_Date"] = WQ_QR_row_cell.value
				WQ_Data[Resp_ID]["RespTime"] = WQ_Data[Resp_ID]["E_Date"] - WQ_Data[Resp_ID]["S_Date"]
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'Name':
				WQ_Data[Resp_ID]["Name"] = WQ_QR_row_cell.value
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'City/Town':
				WQ_Data[Resp_ID]["City"] = WQ_QR_row_cell.value
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'Email Address':
				WQ_Data[Resp_ID]["Email"] = WQ_QR_row_cell.value
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'Phone Number':
				WQ_Data[Resp_ID]["Mobile"] = WQ_QR_row_cell.value
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'Under 18':
				if WQ_QR_row_cell.value != None:
					WQ_Data[Resp_ID]["AgeCat"] = "Cat1_0_18"
					WQ_Data["Summary"]["Cat1_0_18"]["N_Respondents"] += 1
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == '18-24':
				if WQ_QR_row_cell.value != None:
					WQ_Data[Resp_ID]["AgeCat"] = "Cat2_18_24"
					WQ_Data["Summary"]["Cat2_18_24"]["N_Respondents"] += 1
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == '25-34':
				if WQ_QR_row_cell.value != None:
					WQ_Data[Resp_ID]["AgeCat"] = "Cat3_25_34"
					WQ_Data["Summary"]["Cat3_25_34"]["N_Respondents"] += 1
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == '35-44':
				if WQ_QR_row_cell.value != None:
					WQ_Data[Resp_ID]["AgeCat"] = "Cat4_35_44"
					WQ_Data["Summary"]["Cat4_35_44"]["N_Respondents"] += 1
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == '45-54':
				if WQ_QR_row_cell.value != None:
					WQ_Data[Resp_ID]["AgeCat"] = "Cat5_45_54"
					WQ_Data["Summary"]["Cat5_45_54"]["N_Respondents"] += 1
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == '55-64':
				if WQ_QR_row_cell.value != None:
					WQ_Data[Resp_ID]["AgeCat"] = "Cat6_55_64"
					WQ_Data["Summary"]["Cat6_55_64"]["N_Respondents"] += 1
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == '65+':
				if WQ_QR_row_cell.value != None:
					WQ_Data[Resp_ID]["AgeCat"] = "Cat7_65_120"
					WQ_Data["Summary"]["Cat7_65_120"]["N_Respondents"] += 1
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'Female':
				if WQ_QR_row_cell.value != None:
					WQ_Data[Resp_ID]["Gender"] = WQ_QR_row_cell.value
					WQ_Data["Summary"]["Overall"]["N_Females"] += 1
					WQ_Data["Summary"][WQ_Data[Resp_ID]["AgeCat"]]["N_Females"] += 1
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'Male':
				if WQ_QR_row_cell.value != None:
					WQ_Data[Resp_ID]["Gender"] = WQ_QR_row_cell.value
					WQ_Data["Summary"]["Overall"]["N_Males"] += 1
					WQ_Data["Summary"][WQ_Data[Resp_ID]["AgeCat"]]["N_Males"] += 1
			elif WQ_QR_sheet.cell(2, WQ_QR_row_cell.column).value == 'Open-Ended Response':
				if WQ_QR_sheet.cell(1, WQ_QR_row_cell.column).value == 'Your Height (in cm)':
					WQ_Data[Resp_ID]["HeightCm"] = WQ_QR_row_cell.value
				elif WQ_QR_sheet.cell(1, WQ_QR_row_cell.column).value == 'Your Weight (in Kg)':
					WQ_Data[Resp_ID]["WeightKg"] = WQ_QR_row_cell.value
					WQ_Data[Resp_ID]["BMI"] = float(WQ_Data[Resp_ID]["WeightKg"]) / (float(WQ_Data[Resp_ID]["HeightCm"])*0.01)**2
				elif WQ_QR_sheet.cell(1, WQ_QR_row_cell.column).value == 'Enter your Authentication Code. Your Wellness Quotient Score will be sent to you only if there is a valid Authentication Code.':
					WQ_Data[Resp_ID]["AuthCode"] = WQ_QR_row_cell.value



# Calculate the WQ Score from the QuizResponses Excel book
WQ_responses = load_workbook('QuizResponses.xlsx')
WQ_response_sheet = WQ_responses.active

# This is to apply restriction on Points for certain Questions. In order to avoid skewing the WQ score.
#col_list_having_maxvalues = {"CO": 10, "EB": 20} # Q25, Q30
col_list_having_maxvalues = {"CO": 10} # Q25 only

# Calculate individual WQ Score by applying formula
for WQ_response_row in WQ_response_sheet.iter_rows(3,WQ_response_sheet.max_row):
	for WQ_response_row_cell in WQ_response_row:
		# if this is a Points column
		if WQ_response_sheet.cell(2, WQ_response_row_cell.column).value == 'Points':
			final_value = int(WQ_response_row_cell.value.partition('/')[0]) # partition returns a tuple. first part is enough for us
			if col_list_having_maxvalues.get(WQ_response_row_cell.column_letter) != None: # This column has max restriction
				final_value = col_list_having_maxvalues[WQ_response_row_cell.column_letter] if final_value > col_list_having_maxvalues[WQ_response_row_cell.column_letter] else final_value
			WQ_Data[WQ_response_row[0].value]["Temp_Tot"] += final_value # Add all Points
	# Now that all Points are added, apply the formula
	WQ_Data[WQ_response_row[0].value]["Score"] = 100 - (WQ_Data[WQ_response_row[0].value]["Temp_Tot"] / 327 * 100)


#Now that the WQ Scores are ready for all, and other data is captured for each respondent,
# it is time to arrive at complex calculations

# Temp Data Structure for Age Group Calculations & Overall Calculations
TempGroups = {'Cat1_0_18': {'TempScoreList': [],
							   'BMITot': 0,
							   'FormTimeTot': 0,
							   'WQTot': 0
							   },
				 'Cat2_18_24': {'TempScoreList': [],
							   'BMITot': 0,
							   'FormTimeTot': 0,
							   'WQTot': 0
								},
				 'Cat3_25_34': {'TempScoreList': [],
							   'BMITot': 0,
							   'FormTimeTot': 0,
							   'WQTot': 0
								},
				 'Cat4_35_44': {'TempScoreList': [],
							   'BMITot': 0,
							   'FormTimeTot': 0,
							   'WQTot': 0
								},
				 'Cat5_45_54': {'TempScoreList': [],
							   'BMITot': 0,
							   'FormTimeTot': 0,
							   'WQTot': 0
								},
				 'Cat6_55_64': {'TempScoreList': [],
							   'BMITot': 0,
							   'FormTimeTot': 0,
							   'WQTot': 0
								},
				 'Cat7_65_120': {'TempScoreList': [],
							   'BMITot': 0,
							   'FormTimeTot': 0,
							   'WQTot': 0
								 },
				 'Overall': {'TempScoreList': [],
							   'BMITot': 0,
							   'FormTimeTot': 0,
							   'WQTot': 0
								 }
				 }


# For each respondent
for val in WQ_Data.items():
	if val[0] == 'Summary':
		continue

	TempGroups["Overall"]["TempScoreList"].append(val[1]["Score"])
	TempGroups["Overall"]["BMITot"] += val[1]["BMI"]
	TempGroups["Overall"]["FormTimeTot"] += val[1]["RespTime"].total_seconds()
	TempGroups["Overall"]["WQTot"] += val[1]["Score"]

	TempGroups[val[1]["AgeCat"]]["TempScoreList"].append(val[1]["Score"])
	TempGroups[val[1]["AgeCat"]]["BMITot"] += val[1]["BMI"]
	TempGroups[val[1]["AgeCat"]]["FormTimeTot"] += val[1]["RespTime"].total_seconds()
	TempGroups[val[1]["AgeCat"]]["WQTot"] += val[1]["Score"]


# For each respondent - second pass for Percentile
for val in WQ_Data.items():
	if val[0] == 'Summary':
		continue
	WQ_Data[val[0]]["O_Perctl"] = stats.percentileofscore(TempGroups["Overall"]["TempScoreList"], WQ_Data[val[0]]["Score"])
	WQ_Data[val[0]]["AG_Perctl"] = stats.percentileofscore(TempGroups[WQ_Data[val[0]]["AgeCat"]]["TempScoreList"], WQ_Data[val[0]]["Score"])


# Updating Age Group Summary & Overall Summary
for ageG in TempGroups.items():
	if WQ_Data["Summary"][ageG[0]]["N_Respondents"] == 0:
		continue
	WQ_Data["Summary"][ageG[0]]["Median"] = statistics.median(ageG[1]["TempScoreList"])
	WQ_Data["Summary"][ageG[0]]["Avg_BMI"] = ageG[1]["BMITot"] / WQ_Data["Summary"][ageG[0]]["N_Respondents"]
	WQ_Data["Summary"][ageG[0]]["Avg_Form_Time"] = ageG[1]["FormTimeTot"] / WQ_Data["Summary"][ageG[0]]["N_Respondents"]
	WQ_Data["Summary"][ageG[0]]["Avg_WQ_Score"] = ageG[1]["WQTot"] / WQ_Data["Summary"][ageG[0]]["N_Respondents"]

f = open('Results.Out', 'w')
pp_on_file = pprint.PrettyPrinter(indent=4, stream=f)
pp_on_file.pprint(WQ_Data)
f.close

# Now that all the calculations are made and values are available, let us create the individual Respondent's Report.

# Anon function to arrive at the ordinal number notation. For e.g., 1st, 2nd, 3rd, etc...
ordinal = lambda n: "%d%s" % (n,"tsnrhtdd"[(math.floor(n/10)%10!=1)*(n%10<4)*n%10::4])

# For each respondent
for val in WQ_Data.items():
	if val[0] == 'Summary':
		continue

	dirName = "./Results/"+str(WQ_Data[val[0]]['AuthCode'])
	try:
		os.makedirs(dirName)
		print("Directory ", dirName, " Created ")
	except FileExistsError:
		print("Directory ", dirName, " already exists")

	filename = dirName+'/file'+str(val[0])+'.html'
	f = open(filename, 'w')

	# WQ Range
	healthRange = 'NEEDS ATTENTION'
	healthColor = '#FF0000'
	if round(WQ_Data[val[0]]['Score']) > 85:
		healthRange = 'HEALTHY'
		healthColor = '#006400'
	elif round(WQ_Data[val[0]]['Score']) > 70:
		healthRange = 'NEEDS PREVENTION'
		healthColor = '#e8a310'

	# Creating the HTML content
	message = '''<html><head><title>Wellness Quotient</title></head>
	<body>
	<br>
	<table align="center" cellspacing="0" cellpadding="0">
		<td width="310"><img src="../../logo.png"></td>
		<td width="700"><h1 style="color:black;" style="font-family:verdana;">Wellness Quotient Report for '''+\
	str(WQ_Data[val[0]]['Name']).title()+'''</h1></td>
	</table>
	<br>
	<style>
	a{
	text-decoration : none;
	color : '''+healthColor+''';
	}
	</style>
	<style>
	* {box-sizing: border-box}

	.container {
	width: 90%;
	background-color: #C0C0C0;
	}

	.skills {
	text-align: right;
	padding-top: 2px;
	padding-bottom: 2px;
	color: white;
	}

	.wqrange {width: '''+str(round(WQ_Data[val[0]]['Score']))+'''%; background-color: '''+healthColor+''';}
	.agerange {width: '''+str(round(WQ_Data[val[0]]['AG_Perctl']))+'''%; background-color: #2196F3;}
	</style>
	<table align="center" cellspacing="0" cellpadding="5">
		<tr><th align="center" background="../../table_background.PNG" style="background-repeat:no-repeat;" rowspan="2" height="100" width="355"><b>
			<font size="5">Your Wellness Quotient is</font></th>
				<td width="250" align="center"><b><font size="4">Your WQ places you in the <a>"'''+healthRange+'''"</a> range</font></td></tr>
			
		<tr><td width="300"><div class="container">
			<div class="skills wqrange"><b>'''+str(round(WQ_Data[val[0]]['Score']))+'''</b></div></td></tr>
		
		<tr><th align="center" background="../../table_background1.PNG" style="background-repeat:no-repeat;" rowspan="2" height="108" width="150"><b><b><font size="9">'''+\
			str(round(WQ_Data[val[0]]['Score']))+'''</font></th>
				<td width="700" align="center"><b><font size="4">Your WQ is at the '''+\
				ordinal(round(WQ_Data[val[0]]['AG_Perctl']))+''' Percentile among people of your age group</font></td></tr>
		
		<tr><td width="300"><div class="container">
			<div class="skills agerange"><b>'''+str(round(WQ_Data[val[0]]['AG_Perctl']))+'''</b></div></td></tr>
			<td width="300"><br><br><br><b><p id="date"></p></b></td>
		<tr></tr>
			<td width="300"><p><font size="3"><b>Report valid for 6 months</b></font></p></td>
	</table>
	<br><br><br>	 
	<script>
		n =  new Date();
    	y = n.getFullYear();
    	m = n.getMonth() + 1;
    	d = n.getDate();
    	document.getElementById("date").innerHTML = ("Report Generated on " + d + "/" + m + "/" + y).fontsize("3");
	</script>    
	</body>	</html>'''

	f.write(message)
	f.close()

	# Converting the HTML to JPEG
	jpegfilename = dirName+'/file'+str(val[0])+'.jpg'
	if sys.platform == 'win32':
		config = imgkit.config(wkhtmltoimage='C:\Program Files\wkhtmltopdf\\bin\wkhtmltoimage.exe')
	else:
		config = imgkit.config(wkhtmltoimage='/usr/local/bin/wkhtmltoimage')

	imgkit.from_file(filename, jpegfilename, config=config)
